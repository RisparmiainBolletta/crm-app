# -*- coding: utf-8 -*-
# Indica che il file utilizza la codifica UTF-8, utile per gestire caratteri speciali (accenti, simboli, ecc.)

# === IMPORTAZIONI ===
from flask import Flask, render_template, request, redirect, url_for, session, jsonify
###from flask import Flask, request, jsonify, send_from_directory, session, redirect
# Flask: framework web principale
# request: per ricevere dati da POST/GET
# jsonify: per restituire dati in formato JSON
# send_from_directory: per servire file statici
# session: per gestire sessioni utente
# redirect: per reindirizzare l'utente
#from flask import request
from datetime import datetime, date, timedelta

import gspread  # Libreria per interfacciarsi con Google Sheets
from oauth2client.service_account import ServiceAccountCredentials  # Autenticazione con vecchia API
from flask_cors import CORS  # Per abilitare le richieste CORS (Cross-Origin Resource Sharing)
import os  # Per accedere a variabili d’ambiente e funzionalità OS

from google.oauth2 import service_account  # Autenticazione moderna con Google
from googleapiclient.discovery import build  # Per costruire client per API Google (es. Drive)
from googleapiclient.http import MediaFileUpload  # Per caricare file su Google Drive
import tempfile  # Per gestire file temporanei
import json  # Per lavorare con oggetti JSON

# === CONFIGURAZIONE FLASK E CORS ===
app = Flask(__name__) #, static_folder="templates")  # Istanzia l'app Flask e imposta la cartella dei file statici
app.secret_key = 'supersegreto'  # Chiave segreta per gestire le sessioni Flask
CORS(app)  # Abilita CORS su tutte le route (utile per frontend su dominio diverso)

# === 1. CONNESSIONE A GOOGLE SHEETS ===
scope = [
    "https://spreadsheets.google.com/feeds",  # Accesso ai feed di Google Sheets
    "https://www.googleapis.com/auth/drive"   # Accesso completo a Google Drive (necessario per allegati)
]

# Carica le credenziali da una variabile d'ambiente che contiene il JSON delle credenziali
credentials_dict = json.loads(os.environ['GOOGLE_APPLICATION_CREDENTIALS_JSON'])

# Autenticazione per accedere a Google Sheets
###creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope) ###?
creds = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
client = gspread.authorize(creds)  # Client gspread autenticato

# === 2. CONNESSIONE A GOOGLE DRIVE ===

# Ricarica le stesse credenziali per Google Drive
drive_credentials_dict = json.loads(os.environ['GOOGLE_APPLICATION_CREDENTIALS_JSON'])

# Usa la nuova libreria google.oauth2 per autenticarsi con Drive
drive_creds = service_account.Credentials.from_service_account_info(
    drive_credentials_dict,
    scopes=['https://www.googleapis.com/auth/drive']  # Scope necessario per accedere e gestire file su Drive
)
###drive_creds = service_account.Credentials.from_service_account_file(        ###?
###    'credentials.json', scopes=['https://www.googleapis.com/auth/drive']    ###?
###)                                                                           ###?

# Crea il client per l'API Drive
drive_service = build('drive', 'v3', credentials=drive_creds)

# === CONFIGURAZIONE CARTELLA PRINCIPALE ===
MAIN_FOLDER_NAME = "CRM_Documenti"  # Nome della cartella principale su Google Drive dove salvare gli allegati
main_folder_id = None  # Verrà assegnato l’ID della cartella

# Verifica se esiste già una cartella con quel nome (non eliminata)
results = drive_service.files().list(
    q=f"name='{MAIN_FOLDER_NAME}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
    spaces='drive',
    fields="files(id, name)"
).execute()
folders = results.get('files', [])

# Se esiste, usa quella; altrimenti la crea
if folders:
    main_folder_id = folders[0]['id']
else:
    file_metadata = {
        'name': MAIN_FOLDER_NAME,
        'mimeType': 'application/vnd.google-apps.folder'
    }
    folder_created = drive_service.files().create(body=file_metadata, fields='id').execute()
    main_folder_id = folder_created.get('id')

# === 3. APERTURA FOGLI GOOGLE SHEETS ===

# Apre il file "CRM_Database" e accede ai vari fogli di lavoro
clienti_sheet = client.open("CRM_Database").worksheet("Clienti")            # Foglio dei clienti
agenti_sheet = client.open("CRM_Database").worksheet("Agenti")              # Foglio degli agenti
interazioni_sheet = client.open("CRM_Database").worksheet("Interazioni")    # Foglio delle interazioni
filelog_sheet = client.open("CRM_Database").worksheet("File_Allegati")      # Foglio con log dei file/documenti
impostazioni_sheet = client.open("CRM_Database").worksheet("Impostazioni")  # Foglio per dropdown dinamici

# -------------------------------------------------------------------
#  A) FUNZIONI DI LOGIN / LOGOUT
# -------------------------------------------------------------------

# Definisce una route POST per il login. Quando l'utente invia le credenziali, questa funzione viene chiamata.

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # POST da frontend con fetch JSON
        if request.is_json:
            data = request.get_json()
            codice = data.get("codice")
            password = data.get("password")
        else:
            # POST da form HTML
            codice = request.form.get("codice")
            password = request.form.get("password")

        agenti = agenti_sheet.get_all_records()
        for agente in agenti:
            if agente['Codice_Agente'] == codice and agente['Password'] == password:
                session['agente'] = codice
                session['ruolo'] = agente.get('Ruolo', 'agente')  # default = agente
                # Risposta JSON se arriva da JS
                if request.is_json:
                    return jsonify({"message": "Login riuscito", "ruolo": session['ruolo']}), 200
                # Oppure redirect se arriva da form
                return redirect("/")

       # Se login fallisce
        if request.is_json:
            return jsonify({"message": "Credenziali errate"}), 401
        else:
            return render_template("login.html", errore="Credenziali errate")

    # Se metodo GET, mostra la pagina login
    return render_template("login.html")


# Definisce la route GET per il logout.
@app.route("/logout")
def logout():
    # Rimuove il codice agente dalla sessione se presente.
    session.pop('agente', None)
    session.pop('ruolo', None)
    # Reindirizza l’utente alla home page dopo il logout.
    return redirect("/")

# -------- NOME AGENTE ---------------
# Definisce una route GET che restituisce i dati dell'agente loggato, da mostrare nel frontend (es. "Mario Rossi - AG01").

@app.route("/dati-agente", methods=["GET"])
def dati_agente():
    # Verifica se l'agente è loggato, altrimenti ritorna errore HTTP 401.
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    # Recupera il codice agente dalla sessione e tutti i record dal foglio "Agenti".
    codice = session['agente']
    ruolo = session.get("ruolo", "agente")

    agenti = agenti_sheet.get_all_records()

    # Cerca l’agente nel foglio e restituisce: Il nome completo - Il codice
    for agente in agenti:
        if agente['Codice_Agente'] == codice:
            return jsonify({
                "nome_completo": agente.get("Nome", ""),
                "codice": codice,
                "ruolo": ruolo
            })
    # Se è admin, restituisci un profilo fittizio
    if ruolo == "admin":
        return jsonify({
            "nome_completo": "Amministratore",
            "codice": "ADMIN",
            "ruolo": "admin"
        })

    return jsonify({"message": "Utente non trovato"}), 404

# ----------------------
# ---  RUOLI AGENTI ----
# ----------------------
@app.route("/ruoli-agenti", methods=["GET"])
def get_ruoli_agenti():
    agenti = agenti_sheet.get_all_records()
    mappa = {a["Codice_Agente"]: a.get("Ruolo", "agente").lower() for a in agenti}
    return jsonify(mappa)


# -------------------------------------------------------------------
#  B) GESTIONE CLIENTI
# -------------------------------------------------------------------

@app.route("/clienti-admin")
def get_clienti_admin():
    if 'agente' not in session or session.get("ruolo") != "admin":
        return jsonify({"message": "Non autorizzato"}), 401

    records = clienti_sheet.get_all_records()
    return jsonify(records)  # ← Ordine inverso


@app.route("/clienti", methods=["GET"])
def get_clienti():
    codice_agente = session.get('agente')
    if not codice_agente:
        return jsonify({"message": "Non autenticato"}), 401
    records = clienti_sheet.get_all_records()
    # Ogni agente vede solo i suoi
    clienti_filtrati = [c for c in records if str(c.get("Agente")) == codice_agente][::-1]
    return jsonify(clienti_filtrati)

@app.route("/clienti", methods=["POST"])
def add_cliente():
    # Aggiunta nuovo cliente (ID generato auto)
    codice_agente = session.get('agente')
    if not codice_agente:
        return jsonify({"message": "Non autenticato"}), 401

    data = request.json
    records = clienti_sheet.get_all_records()
    clienti_agente = [c for c in records if str(c.get("Agente")) == codice_agente]

    # Trova ID più alto esistente per quell'agente
    numeri_usati = []
    for cliente in clienti_agente:
        id_cliente = cliente.get("ID_Cliente", "")
        if id_cliente.startswith(codice_agente + "-"):
            parte_num = id_cliente.split("-")[-1]
            if parte_num.isdigit():
                numeri_usati.append(int(parte_num))
    prossimo_numero = max(numeri_usati) + 1 if numeri_usati else 1
    nuovo_id = f"{codice_agente}-{prossimo_numero:04d}"

    data_oggi = datetime.today().strftime("%d/%m/%Y")

    new_row = [
        nuovo_id,                    
        data.get("Nome"),           
        data.get("Categoria"),
        data.get("Email"),          
        data.get("Telefono"),
        data.get("Citta"),
        data.get("Provincia"),
        data.get("Stato"),          
        "",                          
        data.get("POD_PDR"),
        data.get("Settore"),
        data.get("Nuovo_Fornitore"),
        data.get("Codice_Fiscale"),
        data.get("Partita_IVA"),
        codice_agente              
    ]
    clienti_sheet.append_row(new_row)
    return jsonify({"message": f"Cliente aggiunto con ID {nuovo_id}"}), 201

@app.route("/clienti/<id_cliente>", methods=["PUT"])
def aggiorna_cliente(id_cliente):
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    data = request.json
    agente = session['agente']
    tutti = clienti_sheet.get_all_records()

    for idx, cliente in enumerate(tutti):
        if cliente["ID_Cliente"] == id_cliente and cliente["Agente"] == agente:
            riga_excel = idx + 2
            stato_precedente = cliente.get("Stato", "").strip().lower()
            stato_nuovo = data.get("Stato", "").strip().lower()

            # 🔁 Aggiorna campi
            clienti_sheet.update(f"B{riga_excel}", [[data["Nome"]]])
            clienti_sheet.update(f"C{riga_excel}", [[data["Categoria"]]])
            clienti_sheet.update(f"D{riga_excel}", [[data["Email"]]])
            clienti_sheet.update(f"E{riga_excel}", [[data["Telefono"]]])
            clienti_sheet.update(f"F{riga_excel}", [[data["Città"]]])
            clienti_sheet.update(f"G{riga_excel}", [[data["Provincia"]]])
            clienti_sheet.update(f"H{riga_excel}", [[data["Stato"]]])
            clienti_sheet.update(f"J{riga_excel}", [[data["POD_PDR"]]])
            clienti_sheet.update(f"K{riga_excel}", [[data["Settore"]]])
            clienti_sheet.update(f"L{riga_excel}", [[data["Nuovo_Fornitore"]]])
            clienti_sheet.update(f"M{riga_excel}", [[data["Codice_Fiscale"]]])
            clienti_sheet.update(f"N{riga_excel}", [[data["Partita_IVA"]]])

            # 🔁 Verifica stato "Da comparare"
            if stato_precedente != stato_nuovo and ("da comparare" in [stato_precedente, stato_nuovo]):
                from requests import post
                try:
                    post("http://127.0.0.1:5000/verifica-clienti-comparare", json={"agente": agente})
                    print(f"🔁 Verifica sincronizzazione file comparazioni per agente {agente}")
                except Exception as e:
                    print("⚠️ Errore nella verifica-clienti-comparare:", e)

            return jsonify({"message": "Cliente aggiornato correttamente"}), 200

    return jsonify({"message": "Cliente non trovato"}), 404


@app.route("/admin/clienti/<id_cliente>", methods=["PUT"])
def admin_modifica_cliente(id_cliente):
    if 'agente' not in session or session.get("ruolo") != "admin":
        return jsonify({"message": "Non autorizzato"}), 401

    data = request.json
    tutti = clienti_sheet.get_all_records()

    for idx, cliente in enumerate(tutti):
        if cliente["ID_Cliente"] == id_cliente:
            riga_excel = idx + 2  # +2 = salta intestazione

            try:
                clienti_sheet.update(f"B{riga_excel}", [[data["Nome"]]])
                clienti_sheet.update(f"C{riga_excel}", [[data["Categoria"]]])
                clienti_sheet.update(f"D{riga_excel}", [[data["Email"]]])
                clienti_sheet.update(f"E{riga_excel}", [[data["Telefono"]]])
                clienti_sheet.update(f"F{riga_excel}", [[data["Città"]]])
                clienti_sheet.update(f"G{riga_excel}", [[data["Provincia"]]])
                clienti_sheet.update(f"H{riga_excel}", [[data["Stato"]]])
                clienti_sheet.update(f"J{riga_excel}", [[data["POD_PDR"]]])
                clienti_sheet.update(f"K{riga_excel}", [[data["Settore"]]])
                clienti_sheet.update(f"L{riga_excel}", [[data["Nuovo_Fornitore"]]])
                clienti_sheet.update(f"M{riga_excel}", [[data["Codice_Fiscale"]]])
                clienti_sheet.update(f"N{riga_excel}", [[data["Partita_IVA"]]])

                return jsonify({"message": "Cliente aggiornato correttamente (admin)"}), 200

            except Exception as e:
                print("❌ Errore durante aggiornamento admin:", e)
                return jsonify({"error": str(e)}), 500

    return jsonify({"message": "Cliente non trovato"}), 404



@app.route("/clienti/<id_cliente>", methods=["DELETE"])
def elimina_cliente(id_cliente):
    if 'agente' not in session and session.get("ruolo", "").strip().lower() != "admin":
        return jsonify({"message": "Non autenticato"}), 401

    tutti = clienti_sheet.get_all_records()
    for idx, cliente in enumerate(tutti):
        if cliente["ID_Cliente"] == id_cliente:
            # Se è un agente, verifica che sia il suo cliente
            if session.get("ruolo", "").strip().lower() != "admin" and cliente["Agente"] != session['agente']:
                continue  # Salta, non può eliminarlo

            riga_excel = idx + 2
            clienti_sheet.delete_rows(riga_excel)

            # 🔁 ELIMINA cartella su Google Drive
            try:
                query = (f"name='{id_cliente}' and mimeType='application/vnd.google-apps.folder' "
                         f"and '{main_folder_id}' in parents and trashed=false")
                results = drive_service.files().list(
                    q=query, spaces='drive', fields="files(id)").execute()
                folders = results.get('files', [])
                if folders:
                    folder_id = folders[0]['id']
                    drive_service.files().delete(fileId=folder_id).execute()
                    print(f"✅ Cartella Drive eliminata per cliente {id_cliente} (ID: {folder_id})")
                else:
                    print(f"ℹ️ Nessuna cartella trovata per cliente {id_cliente}")
            except Exception as e:
                print("⚠️ Errore durante eliminazione cartella Drive:", e)

            # 🔁 ELIMINA righe dal foglio File_Allegati
            try:
                log_records = filelog_sheet.get_all_records()
                righe_da_eliminare = [i + 2 for i, r in enumerate(log_records) if r.get("ID_Cliente") == id_cliente]

                for i in reversed(righe_da_eliminare):
                    filelog_sheet.delete_rows(i)

                print(f"🗑️ Righe eliminate dal log File_Allegati per cliente {id_cliente}: {len(righe_da_eliminare)}")
            except Exception as e:
                print("⚠️ Errore durante eliminazione righe File_Allegati:", e)

            return jsonify({"message": "Cliente e documenti eliminati correttamente"}), 200

    return jsonify({"message": "Cliente non trovato"}), 404


# -------------------------------------------------------------------
#  C) GESTIONE INTERAZIONI
# -------------------------------------------------------------------
@app.route("/interazioni/<id_cliente>", methods=["GET"])
def get_interazioni(id_cliente):
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    codice_agente = session.get('agente', '').strip().upper()
    ruolo = session.get('ruolo', 'agente').strip().lower()

    records = interazioni_sheet.get_all_records()
    agenti = agenti_sheet.get_all_records()

    # ✅ Crea una lista di codici agenti con ruolo = admin
    codici_admin = [
        a['Codice_Agente'].strip().upper()
        for a in agenti
        if a.get('Ruolo', '').strip().lower() == 'admin'
    ]

    interazioni_cliente = []

    for r in records:
        id_cli = str(r.get("ID_Cliente", "")).strip()
        agente_interazione = str(r.get("Agente", "")).strip().upper()

        if id_cli == id_cliente:
            if ruolo == "admin":
                interazioni_cliente.append(r)  # Admin vede tutto
            else:
                if agente_interazione == codice_agente or agente_interazione in codici_admin:
                    interazioni_cliente.append(r)

    return jsonify(interazioni_cliente)



# ✅ 1. AGGIORNA la route POST /interazioni per inserire 'Letto' = FALSE
@app.route("/interazioni", methods=["POST"])
def aggiungi_interazione():
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    data = request.json
    id_cliente = data.get("ID_Cliente")
    tipo = data.get("Tipo")
    esito = data.get("Esito")
    descrizione = data.get("Descrizione", "").strip()

    if not id_cliente or not tipo or not esito:
        return jsonify({"message": "Campi obbligatori mancanti"}), 400

    tutte = interazioni_sheet.get_all_records()
    nuovo_id = f"I{len(tutte)+1:04d}"

    data_interazione = data.get("Data")
    if not data_interazione:
        data_interazione = datetime.now().strftime("%d/%m/%Y")

    agente = session['agente']

    # Se Admin inserisce interazioni, salviamo come 'ADMIN'
    if isinstance(agente, dict) and agente.get("ruolo", "").strip().lower() == "admin":
        agente_nome = "ADMIN"
    else:
        agente_nome = agente

    nuova_riga = [
        nuovo_id,
        id_cliente,
        agente_nome,
        data_interazione,
        tipo,
        esito,
        descrizione,
        "FALSE"  # 🔔 Letto = FALSE
    ]

    interazioni_sheet.append_row(nuova_riga)

    return jsonify({"message": "Interazione salvata con successo", "ID": nuovo_id})


# ✅ 2. NUOVA route: ritorna tutti gli ID_Cliente con interazioni 'Letto = FALSE'
# ... ma solo quelle non inserite dall'utente attuale
@app.route("/notifiche-interazioni-non-letto", methods=["GET"])
def notifiche_interazioni_non_lette():
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    agente_attuale = session['agente'].strip().upper()
    ruolo = session.get("ruolo", "agente").strip().lower()

    records = interazioni_sheet.get_all_records()
    clienti_con_notifiche = set()

    for r in records:
        letto = str(r.get("Letto", "")).strip().upper()
        agente_interazione = str(r.get("Agente", "")).strip().upper()
        id_cliente = str(r.get("ID_Cliente", "")).strip()

        if letto != "FALSE":
            continue

        # Mostra la notifica solo se l'interazione NON è stata inserita dall'utente stesso
        if agente_interazione != agente_attuale:
            clienti_con_notifiche.add(id_cliente)

    return jsonify(list(clienti_con_notifiche))

# ✅ 3. NUOVA route: segna tutte le interazioni di un cliente come lette
@app.route("/interazioni/segna-letti/<id_cliente>", methods=["POST"])
def segna_interazioni_lette(id_cliente):
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    agente_attuale = session.get("agente", "").strip().upper()
    records = interazioni_sheet.get_all_records()

    aggiornate = 0

    for i, r in enumerate(records):
        id_cli = str(r.get("ID_Cliente", "")).strip()
        agente_interazione = str(r.get("Agente", "")).strip().upper()
        letto = str(r.get("Letto", "")).strip().upper()

        if (
            id_cli == id_cliente
            and letto == "FALSE"
            and agente_interazione != agente_attuale
        ):
            try:
                interazioni_sheet.update_cell(i + 2, 8, "TRUE")  # Colonna 8 = Letto
                aggiornate += 1
            except Exception as e:
                print(f"❌ Errore aggiornamento riga {i+2}: {e}")

    return jsonify({"message": f"Interazioni segnate come lette: {aggiornate}"})



# -------------------------------------------------------------------
#  D) GESTIONE DOCUMENTI
# -------------------------------------------------------------------
@app.route("/upload/<id_cliente>", methods=["POST"])
def upload_file(id_cliente):
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401
    if 'file' not in request.files:
        return jsonify({"message": "Nessun file ricevuto"}), 400

    file = request.files['file']
    filename = file.filename
    codice_agente = session['agente']
    ruolo = request.form.get("ruolo", "agente").lower().strip()
    data_oggi = datetime.today().strftime("%d/%m/%Y")

    # Trova o crea cartella del cliente su Drive
    query = (f"name='{id_cliente}' and mimeType='application/vnd.google-apps.folder' "
             f"and '{main_folder_id}' in parents and trashed=false")
    results = drive_service.files().list(q=query, spaces='drive', fields="files(id, name)").execute()
    folders = results.get('files', [])
    if folders:
        folder_id = folders[0]['id']
    else:
        file_metadata = {
            'name': id_cliente,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [main_folder_id]
        }
        folder = drive_service.files().create(body=file_metadata, fields='id').execute()
        folder_id = folder.get('id')

    # Crea file temporaneo
    import shutil
    temp_fd, temp_path = tempfile.mkstemp()
    with os.fdopen(temp_fd, 'wb') as tmp:
        shutil.copyfileobj(file.stream, tmp)

    # Caricamento su Drive
    media = MediaFileUpload(temp_path)
    file_metadata = {
        'name': filename,
        'parents': [folder_id],
        'mimeType': file.mimetype or 'application/octet-stream'
    }

    uploaded = drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields='id'
    ).execute()

    file_id = uploaded.get("id")

    # 🔓 Rendi il file visibile pubblicamente
    drive_service.permissions().create(
        fileId=file_id,
        body={
            "role": "reader",
            "type": "anyone",
            "allowFileDiscovery": False
        }
    ).execute()

    # Pulizia file temporaneo
    try:
        os.remove(temp_path)
    except Exception as e:
        print("⚠️ Impossibile eliminare il file temporaneo:", e)

    # Log su File_Allegati con ruolo e notifiche incrociate
    try:
        caricato_da = "ADMIN" if ruolo == "admin" else "AGENTE"
        letto = "FALSE"  # nuovo campo unico

        print("✅ Registro file:", file_id, id_cliente, filename, codice_agente, data_oggi)
        filelog_sheet.append_row([
            file_id,            # ID Google Drive
            id_cliente,
            filename,
            codice_agente,
            data_oggi,
            "In attesa",        # stato
            caricato_da,        # "ADMIN" o "AGENTE"
            letto               # 🔁 unico campo
        ])
    except Exception as e:
        print("❌ Errore durante la scrittura su File_Allegati:", str(e))
        return jsonify({"message": "File caricato, ma errore nel log", "error": str(e)}), 500

    return jsonify({"message": "File caricato e registrato correttamente"})


@app.route("/files/<id_cliente>", methods=["GET"])
def get_files(id_cliente):
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    # Trova cartella del cliente
    query = (f"name='{id_cliente}' and mimeType='application/vnd.google-apps.folder' "
             f"and '{main_folder_id}' in parents and trashed=false")
    results = drive_service.files().list(q=query, spaces='drive', fields="files(id, name)").execute()
    folders = results.get('files', [])
    if not folders:
        return jsonify([])

    folder_id = folders[0]['id']
    query = f"'{folder_id}' in parents and trashed=false"
    files = drive_service.files().list(q=query, spaces='drive', 
                                       fields="files(id, name)").execute().get('files', [])

    log_records = filelog_sheet.get_all_records()
    file_links = []
    for f in files:
        stato = "In attesa"
        caricato_da = "AGENTE"
        letto = "TRUE"
        for r in log_records:
            if r['ID_File'] == f['id']:
                stato = r.get('Stato', "In attesa")
                caricato_da = r.get('Caricato_da', "AGENTE")
                letto = r.get('Letto_da_Agente', "TRUE")
                break
        file_links.append({
            'name': f['name'],
            'id': f['id'],
            'url': f"https://drive.google.com/uc?export=view&id={f['id']}",
            'stato': stato,
            'caricato_da': caricato_da,
            'letto_da_agente': letto
        })

    return jsonify(file_links)

@app.route("/file/<file_id>", methods=["DELETE"])                         # Quando viene effettuata una richiesta DELETE all'URL /file/qualcosa, questa funzione viene chiamata. Il file_id è il parametro (ID del file su Google Drive).
def delete_file(file_id):                                                 # ...>
    if 'agente' not in session:                                           # Verifica che l'utente sia autenticato (loggato come agente). Se no, restituisce errore 401.
        return jsonify({"message": "Non autenticato"}), 401               # ...>

        agente = session['agente']
        is_admin = agente.get("ruolo", "").upper() == "ADMIN"  # ✅ Verifica se l'utente è Admin
    
    try:
            agente = session.get('agente', {})
            is_admin = isinstance(agente, dict) and agente.get("ruolo", "").strip().lower() == "admin"
                                                                # Recupera tutti i record del foglio Google "File_Allegati".
            log_records = filelog_sheet.get_all_records()                       # ...>
            for idx, r in enumerate(log_records):                               # Cerca nel foglio il record corrispondente all’ID del file passato nella richiesta. idx serve a calcolare la riga corretta per l'aggiornamento.
                if r['ID_File'].strip() == file_id.strip():                     # ...>
                    stato_corrente = r.get("Stato", "").strip().lower()         # Se il file è stato approvato, non può essere eliminato (controllo di sicurezza). Restituisce errore 403 (forbidden).
                
                    # ⛔ Blocca solo se NON sei admin e il file è approvato
                    if stato_corrente == "approvato" and not is_admin:                                                      # ...
                        return jsonify({"message": "File approvato. Non può essere eliminato."}), 403       # ...>

                    # ✅ Elimina da Drive
                    try:                                                                    # Se non è approvato, elimina fisicamente il file da Google Drive tramite le API.
                        drive_service.files().delete(fileId=file_id).execute()              # ...>
                    except Exception as e:                                                                                  # Se qualcosa va storto con l'API Google Drive, invia un messaggio di errore 500.
                        return jsonify({"message": "Errore durante l'eliminazione da Drive", "error": str(e)}), 500         # ...>

                    # ✅ Aggiorna colonna "Stato" con "ELIMINATO"
                    riga_excel = idx + 2  # intestazione + base 1                               # Se il file viene eliminato con successo, aggiorna la colonna "Stato" (colonna F) nel foglio Google, scrivendo "ELIMINATO" nella riga corrispondente.
                    colonna_stato = 6     # colonna F                                           # ...
                    filelog_sheet.update_cell(riga_excel, colonna_stato, "ELIMINATO")           # ...>

                    return jsonify({"message": "File eliminato e stato aggiornato"}), 200       # Tutto ok → restituisce messaggio di successo.

            return jsonify({"message": "File non trovato nel log"}), 404                        # Nessun file trovato con quell’ID → errore 404.

    except Exception as e:                                                                  # Qualsiasi altro errore generale viene catturato e restituito come errore 500.
        return jsonify({"message": "Errore imprevisto", "error": str(e)}), 500              # ...>


# Questa route non esegue alcun controllo sul ruolo o stato del file: elimina e basta, solo se è raggiunta da chi la conosce (cioè la modale Admin).
@app.route("/admin/elimina-file/<file_id>", methods=["DELETE"])
def delete_file_admin(file_id):
    print("➡️ ADMIN: richiesta di eliminazione per file:", file_id)

    if "agente" not in session:
        return jsonify({"message": "Non autenticato"}), 401

    try:
        log_records = filelog_sheet.get_all_records()
        for idx, r in enumerate(log_records):
            if r.get('ID_File', '').strip() == file_id.strip():
                print(f"🗑 Eliminazione da Google Drive: {file_id}")

                # Elimina da Drive
                try:
                    drive_service.files().delete(fileId=file_id).execute()
                except Exception as e:
                    print("❌ Errore Drive:", e)
                    return jsonify({"message": "Errore durante l'eliminazione da Drive", "error": str(e)}), 500

                # Aggiorna foglio
                filelog_sheet.update_cell(idx + 2, 6, "ELIMINATO")
                return jsonify({"message": "File eliminato (admin)"}), 200

        print("⚠️ File non trovato nel log")
        return jsonify({"message": "File non trovato nel log"}), 404

    except Exception as e:
        print("❌ ERRORE ADMIN DELETE FILE:", e)
        return jsonify({"message": "Errore imprevisto", "error": str(e)}), 500



# ✅ CORRETTO: route a livello principale
@app.route("/conteggio-documenti", methods=["GET"])                                 # Crea una route HTTP GET all’indirizzo /conteggio-documenti.
def conteggio_documenti():                                                          # Definisce la funzione che verrà eseguita quando si visita l’endpoint.
    if 'agente' not in session:                                                     # Verifica che l’agente sia autenticato. Se no, restituisce un errore 401 Unauthorized.
        return jsonify({"message": "Non autenticato"}), 401                         # ...>

    log_records = filelog_sheet.get_all_records()                                   #  Recupera tutti i record dal foglio File_Allegati. Ogni riga corrisponde a un file caricato da un agente o da un admin.
    conteggi = {}                                                                   # Crea un dizionario vuoto dove salveremo quanti file ha ogni cliente.

    for r in log_records:                                                           # Cicla tutte le righe della tabella File_Allegati.
        id_cliente = r.get("ID_Cliente", "").strip()                                # Estrae l'ID del cliente e lo stato del file (es. "In attesa", "Approvato", "ELIMINATO"), rimuovendo spazi e forzando tutto in maiuscolo.
        stato = r.get("Stato", "").strip().upper()                                  # ...>
        if id_cliente and stato != "ELIMINATO":                                     # Considera solo i file non eliminati (quindi quelli validi e visibili).
            conteggi[id_cliente] = conteggi.get(id_cliente, 0) + 1                  # Aggiorna il conteggio per l’ID_Cliente. Se non c’è ancora nel dizionario, parte da 0.

    return jsonify(conteggi)                                                        # Restituisce il dizionario come risposta JSON, utilizzabile in JavaScript per mostrare i numeri accanto al pulsante "📎 Documenti".

@app.route("/notifiche-non-letto", methods=["GET"])
def notifiche_non_lette():
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    ruolo = session.get("ruolo", "agente").lower()
    codice_agente = session.get("agente")
    log_records = filelog_sheet.get_all_records()
    clienti_records = clienti_sheet.get_all_records()

    clienti_agente = [r["ID_Cliente"] for r in clienti_records if r.get("Agente") == codice_agente]
    clienti_con_notifiche = set()

    for r in log_records:
        id_cliente = r.get("ID_Cliente")
        caricato_da = r.get("Caricato_da", "").strip().upper()
        letto = r.get("Letto", "").strip().upper()

        if ruolo != "admin" and id_cliente not in clienti_agente:
            continue

        # Admin vede notifiche se file caricato da AGENTE
        # Agente vede notifiche se file caricato da ADMIN
        if (
            (ruolo == "admin" and caricato_da == "AGENTE" and letto != "TRUE") or
            (ruolo == "agente" and caricato_da == "ADMIN" and letto != "TRUE")
        ):
            clienti_con_notifiche.add(id_cliente)

    return jsonify(list(clienti_con_notifiche))



# -------------------------------------------------------------------
#  D1) UPLOAD FILE DA ADMIN
# -------------------------------------------------------------------
# Visualizza pagina Admin upload
@app.route("/admin-upload")
def pagina_admin_upload():
    return send_from_directory(app.static_folder, "admin_upload.html")


# Admin carica un file
@app.route("/admin-upload/<id_cliente>", methods=["POST"])
def admin_upload_file(id_cliente):
    if 'file' not in request.files:
        return jsonify({"message": "Nessun file ricevuto"}), 400

    file = request.files['file']
    filename = file.filename
    data_oggi = datetime.today().strftime("%d/%m/%Y")

    # Trova ID_Agente del cliente
    clienti = clienti_sheet.get_all_records(value_render_option='UNFORMATTED_VALUE')
    agente_cliente = next((c["Agente"] for c in clienti if c["ID_Cliente"] == id_cliente), None)

    if not agente_cliente:
        return jsonify({"message": "Cliente non trovato"}), 404

    # Trova o crea cartella su Drive
    query = (f"name='{id_cliente}' and mimeType='application/vnd.google-apps.folder' "
             f"and '{main_folder_id}' in parents and trashed=false")
    results = drive_service.files().list(q=query, spaces='drive', fields="files(id, name)").execute()
    folders = results.get('files', [])
    if folders:
        folder_id = folders[0]['id']
    else:
        folder_metadata = {
            'name': id_cliente,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [main_folder_id]
        }
        folder = drive_service.files().create(body=folder_metadata, fields='id').execute()
        folder_id = folder.get('id')

    # Salva file temporaneo
    import shutil
    temp_fd, temp_path = tempfile.mkstemp()
    with os.fdopen(temp_fd, 'wb') as tmp:
        shutil.copyfileobj(file.stream, tmp)

    media = MediaFileUpload(temp_path)
    file_metadata = {'name': filename, 'parents': [folder_id]}
    uploaded = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    file_id = uploaded.get('id')

    # Rendi il file visibile pubblicamente
    drive_service.permissions().create(
        fileId=file_id,
        body={
            "role": "reader",
            "type": "anyone",
            "allowFileDiscovery": False
        }
    ).execute()

    try:
        os.remove(temp_path)
    except Exception as e:
        print("⚠️ Impossibile eliminare il file temporaneo:", e)

    # Log
    try:
        filelog_sheet.append_row([
            file_id,
            id_cliente,
            filename,
            agente_cliente,
            data_oggi,
            "Approvato",       # ✅ Stato
            "ADMIN",           # ✅ Caricato_da
            "FALSE"            # ✅ Letto_da_Agente
        ])
    except Exception as e:
        return jsonify({"message": "File caricato, ma errore nel log", "error": str(e)}), 500

    return jsonify({"message": "File caricato correttamente da Admin"})

# APPROVA FILE
@app.route("/approva-file/<file_id>", methods=["POST"])
def approva_file(file_id):
    if 'agente' not in session or session.get("ruolo") != "admin":
        return jsonify({"message": "Non autorizzato"}), 401

    try:
        log_records = filelog_sheet.get_all_records()
        for idx, r in enumerate(log_records):
            if r['ID_File'].strip() == file_id.strip():
                riga_excel = idx + 2
                filelog_sheet.update_cell(riga_excel, 6, "Approvato")  # Colonna F = Stato
                return jsonify({"message": "File approvato con successo"})
        return jsonify({"message": "File non trovato"}), 404

    except Exception as e:
        return jsonify({"message": "Errore imprevisto", "error": str(e)}), 500


# -------------------------------------------------------------------
#  E) GESTIONE STATI, TIPI, ESITI (Impostazioni)
# -------------------------------------------------------------------
@app.route("/stati-cliente", methods=["GET"])
def get_stati_cliente():
    try:
        valori = impostazioni_sheet.col_values(3)
        stati = [v for v in valori if v.lower() != "stato_cliente" and v.strip() != ""]
        return jsonify(stati)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/tipi-interazione", methods=["GET"])
def get_tipi_interazione():
    try:
        valori = impostazioni_sheet.col_values(1)
        tipi = [v for v in valori if v.lower() != "tipo_interazione" and v.strip() != ""]
        return jsonify(tipi)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/esiti-interazione", methods=["GET"])
def get_esiti_interazione():
    try:
        valori = impostazioni_sheet.col_values(2)
        esiti = [v for v in valori if v.lower() != "esito_interazione" and v.strip() != ""]
        return jsonify(esiti)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/categorie", methods=["GET"])
def get_categorie():
    try:
        valori = impostazioni_sheet.col_values(4)
        categorie = [v for v in valori if v.lower() != "categoria" and v.strip() != ""]
        return jsonify(categorie)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/settori", methods=["GET"])
def get_settori():
    try:
        valori = impostazioni_sheet.col_values(5)
        settori = [v for v in valori if v.lower() != "settore" and v.strip() != ""]
        return jsonify(settori)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Viene chiamata quando uno specifico cliente cambia stato in "Da comparare"
@app.route("/sincronizza-da-comparare/<id_cliente>", methods=["POST"])
def sincronizza_cliente_da_comparare(id_cliente):
    from openpyxl import Workbook, load_workbook
    from io import BytesIO
    from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    try:
        stato_corrente = request.json.get("Stato", "").strip().lower()
        codice_agente = session["agente"]

        clienti = clienti_sheet.get_all_records()
        cliente = next((c for c in clienti if c.get("ID_Cliente") == id_cliente and c.get("Agente") == codice_agente), None)
        if not cliente:
            return jsonify({"message": "Cliente non trovato"}), 404

        nome_file = f"clienti_da_comparare_{codice_agente}.xlsx"
        nome_cartella = "Da_Comparare"

        # Recupera o crea cartella
        query = (
            f"name = '{nome_cartella}' and mimeType = 'application/vnd.google-apps.folder' "
            f"and '{main_folder_id}' in parents and trashed = false"
        )
        results = drive_service.files().list(q=query, spaces='drive', fields="files(id)").execute()
        folders = results.get("files", [])

        if folders:
            sotto_cartella_id = folders[0]["id"]
        else:
            file_metadata = {
                "name": nome_cartella,
                "mimeType": "application/vnd.google-apps.folder",
                "parents": [main_folder_id]
            }
            folder = drive_service.files().create(body=file_metadata, fields="id").execute()
            sotto_cartella_id = folder["id"]

        # Cerca file
        query = (
            f"name = '{nome_file}' and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
            f"and '{sotto_cartella_id}' in parents and trashed = false"
        )
        results = drive_service.files().list(q=query, spaces='drive', fields="files(id)").execute()
        files = results.get("files", [])

        file_id = None
        intestazioni = list(cliente.keys())

        if files:
            file_id = files[0]["id"]
            request_drive = drive_service.files().get_media(fileId=file_id)
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request_drive)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            fh.seek(0)

            wb = load_workbook(fh)
            ws = wb.active
            header_row = [cell.value for cell in ws[1]]
            if not header_row or "ID_Cliente" not in header_row:
                ws.delete_rows(1, ws.max_row)
                ws.append(intestazioni)
            else:
                intestazioni = header_row
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(intestazioni)

        id_col = intestazioni.index("ID_Cliente")
        righe = list(ws.iter_rows(min_row=2, values_only=True))
        nuovi_dati = []
        gia_presente = False

        for r in righe:
            if str(r[id_col]) == id_cliente:
                gia_presente = True
                if stato_corrente != "da comparare":
                    continue
            nuovi_dati.append(r)

        if not gia_presente and stato_corrente == "da comparare":
            nuova_riga = [cliente.get(col, "") for col in intestazioni]
            nuovi_dati.append(tuple(nuova_riga))
            print(f"➕ Aggiunto {id_cliente} a {nome_file}")
        elif gia_presente and stato_corrente != "da comparare":
            print(f"➖ Rimosso {id_cliente} da {nome_file}")

        ws.delete_rows(2, ws.max_row)
        for r in nuovi_dati:
            ws.append(r)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

        if file_id:
            drive_service.files().update(fileId=file_id, media_body=media).execute()
        else:
            file_metadata = {
                'name': nome_file,
                'parents': [sotto_cartella_id],
                'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()

        return jsonify({"message": f"File sincronizzato per {id_cliente}"}), 200

    except Exception as e:
        print(f"❌ Errore nella sincronizzazione cliente: {e}")
        return jsonify({"message": "Errore nella sincronizzazione", "error": str(e)}), 500



@app.route("/verifica-clienti-comparare", methods=["POST"])
def verifica_clienti_da_comparare():
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    try:
        from openpyxl import Workbook, load_workbook
        from io import BytesIO
        from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

        agente = session["agente"]
        nome_file = f"clienti_da_comparare_{agente}.xlsx"
        nome_cartella = "Da_Comparare"

        # 1. Recupera o crea la cartella "Da_Comparare" dentro CRM_Documenti
        query = (
            f"name = '{nome_cartella}' and mimeType = 'application/vnd.google-apps.folder' "
            f"and '{main_folder_id}' in parents and trashed = false"
        )
        results = drive_service.files().list(q=query, spaces='drive', fields="files(id)").execute()
        folders = results.get("files", [])

        if folders:
            sotto_cartella_id = folders[0]["id"]
        else:
            file_metadata = {
                "name": nome_cartella,
                "mimeType": "application/vnd.google-apps.folder",
                "parents": [main_folder_id]
            }
            folder = drive_service.files().create(body=file_metadata, fields="id").execute()
            sotto_cartella_id = folder["id"]

        # 2. Filtra clienti dell’agente con stato "Da comparare"
        clienti = clienti_sheet.get_all_records()
        clienti_da_comparare = [c for c in clienti if c.get("Agente") == agente and c.get("Stato", "").strip().lower() == "da comparare"]

        # 3. Cerca il file dell’agente su Drive
        query = (
            f"name = '{nome_file}' and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
            f"and '{sotto_cartella_id}' in parents and trashed = false"
        )
        results = drive_service.files().list(q=query, spaces='drive', fields="files(id, name)").execute()
        files = results.get('files', [])

        file_id = None
        wb = Workbook()
        ws = wb.active
        intestazioni = list(clienti[0].keys()) if clienti else []

        if files:
            file_id = files[0]["id"]
            request_drive = drive_service.files().get_media(fileId=file_id)
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request_drive)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            fh.seek(0)
            wb = load_workbook(fh)
            ws = wb.active
            intestazioni = [cell.value for cell in ws[1]]
        else:
            ws.append(intestazioni)

        # 4. Rimuove righe non più "Da comparare"
        id_col = intestazioni.index("ID_Cliente")
        righe = list(ws.iter_rows(min_row=2, values_only=True))
        da_comparare_ids = [c["ID_Cliente"] for c in clienti_da_comparare]
        righe_valide = [r for r in righe if r[id_col] in da_comparare_ids]

        ws.delete_rows(2, ws.max_row)
        for riga in righe_valide:
            ws.append(riga)

        # 5. Aggiunge nuovi clienti
        gia_presenti = [r[id_col] for r in righe_valide]
        for cliente in clienti_da_comparare:
            if cliente["ID_Cliente"] not in gia_presenti:
                ws.append([cliente.get(col, "") for col in intestazioni])

        # 6. Salva il file aggiornato su Drive
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)

        if file_id:
            drive_service.files().update(fileId=file_id, media_body=media).execute()
        else:
            file_metadata = {
                'name': nome_file,
                'parents': [sotto_cartella_id],
                'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()

        return jsonify({"message": f"Verifica completata per {nome_file}"}), 200

    except Exception as e:
        print(f"❌ Errore verifica comparazione: {e}")
        return jsonify({"message": "Errore durante la verifica", "error": str(e)}), 500



# -------------------------------------------------------------------
#  F) Quando l’agente apre i documenti → aggiorna
# -------------------------------------------------------------------
@app.route("/files/segna-letti/<id_cliente>", methods=["POST"])
def segna_letti(id_cliente):
    if "agente" not in session:
        return jsonify({"message": "Non autenticato"}), 401

    ruolo = session.get("ruolo", "agente").lower()
    logs = filelog_sheet.get_all_records()

    for idx, row in enumerate(logs):
        if row.get("ID_Cliente") != id_cliente:
            continue

        caricato_da = row.get("Caricato_da", "").strip().upper()
        letto = row.get("Letto", "").strip().upper()

        # Evita di marcare come letto i file caricati dallo stesso ruolo
        if (
            (ruolo == "admin" and caricato_da == "AGENTE") or
            (ruolo == "agente" and caricato_da == "ADMIN")
        ):
            if letto != "TRUE":
                filelog_sheet.update_cell(idx + 2, 8, "TRUE")  # colonna "Letto"

    return jsonify({"message": "Notifiche segnate come lette"})




# -------------------------------------------------------------------
#  F) ROUTE PRINCIPALI DI AVVIO
# -------------------------------------------------------------------

# Controlla che l’agente sia loggato
@app.route('/')
def home():
    if 'agente' not in session:
        return redirect("/login") # Altrimenti reindirizza al login
    
    # Se è un admin, lo rimanda alla pagina admin
    if session.get("ruolo") == "admin":
        return redirect("/admin")

    # Altrimenti carica i clienti dell'agente
    codice_agente = session['agente']
    records = clienti_sheet.get_all_records()
    clienti = [c for c in records if str(c.get("Agente")) == codice_agente][::-1]

    # Carica index.html con i clienti filtrati
    return render_template('index.html', clienti=clienti)

# Verifica che ci sia un agente loggato e che abbia ruolo admin
@app.route("/admin")
def pagina_admin():
    # Verifica che l'utente sia loggato e sia un admin
    if 'agente' not in session or session.get('ruolo') != 'admin':
        return redirect("/login")  # Reindirizza se non autorizzato
    
    # Mostra la pagina admin.html per l’interfaccia dedicata
    return render_template("admin.html")



# -------------------------------------------------------------------
#  G) PROVVIGIONI
# -------------------------------------------------------------------
@app.route("/provvigioni", methods=["GET"])
def get_provvigioni():
    if 'agente' not in session:
        return jsonify({"message": "Non autenticato"}), 401

    codice_agente = session['agente']
    clienti = clienti_sheet.get_all_records()
    clienti = clienti_sheet.get_all_records(value_render_option='UNFORMATTED_VALUE')

    # Filtro solo i clienti dell'agente con provvigione valorizzata
    risultati = []
    for cliente in clienti:
        if cliente.get("Agente") == codice_agente and cliente.get("Provvigione"):
            risultato = {
                #"Competenza": datetime.today().strftime("%m/%Y"),
                "Competenza": cliente.get("Competenza"),
                "ID_Cliente": cliente.get("ID_Cliente"),
                "Nome": cliente.get("Nome"),
                "Categoria": cliente.get("Categoria"),
                "Stato": cliente.get("Stato"),
                "Settore": cliente.get("Settore"),
                "Nuovo_Fornitore": cliente.get("Nuovo_Fornitore"),
                "Provvigione": cliente.get("Provvigione"),
                "Modificabile": cliente.get("Stato") != "Contratto ATTIVATO"
            }
            risultati.append(risultato)

    return jsonify(risultati)

@app.route("/pagina-provvigioni")
def pagina_provvigioni():
    return render_template("provvigioni.html")

# --- Provvigioni ADMIN ---
@app.route("/provvigioni-admin")
def pagina_provvigioni_admin():
    if 'agente' not in session or session.get("ruolo", "").strip().lower() != "admin":
        return redirect("/login")

    return render_template("provvigioni-admin.html")

@app.route("/provvigioni-tutte", methods=["GET"])
def get_provvigioni_admin():
    if session.get("ruolo", "").strip().lower() != "admin":
        return jsonify({"message": "Non autorizzato"}), 401

    clienti = clienti_sheet.get_all_records()
    clienti = clienti_sheet.get_all_records(value_render_option='UNFORMATTED_VALUE')
    risultati = []

    for cliente in clienti:
        #if cliente.get("Provvigione"):
            risultato = {
                #"Competenza": datetime.today().strftime("%m/%Y"),
                "Competenza": cliente.get("Competenza"),
                "ID_Cliente": cliente.get("ID_Cliente"),
                "Nome": cliente.get("Nome"),
                "Categoria": cliente.get("Categoria"),
                "Stato": cliente.get("Stato"),
                "Settore": cliente.get("Settore"),
                "Nuovo_Fornitore": cliente.get("Nuovo_Fornitore"),
                "Provvigione": cliente.get("Provvigione"),
                "Agente": cliente.get("Agente"),
                "Modificabile": cliente.get("Stato") != "Contratto ATTIVATO"
            }
            risultati.append(risultato)

    return jsonify(risultati)


# Salva le modifiche apportate nella tabella Provvigioni

@app.route("/admin/provvigioni/<id_cliente>", methods=["PUT"])
def aggiorna_provvigione_admin(id_cliente):
    if session.get("ruolo", "").strip().lower() != "admin":
        return jsonify({"message": "Non autorizzato"}), 401

    dati = request.get_json()
    clienti = clienti_sheet.get_all_records()

    for idx, cliente in enumerate(clienti):
        if cliente["ID_Cliente"] == id_cliente:
            riga = idx + 2  # +2 = salta intestazione (base 1)
            try:
                clienti_sheet.update(f"P{riga}", [[dati["Competenza"]]])  # colonna I: Competenza
                clienti_sheet.update(f"I{riga}", [[dati["Provvigione"]]])  # colonna O: Provvigione
                return jsonify({"message": "Dati aggiornati"}), 200
            except Exception as e:
                print("❌ Errore aggiornamento:", e)
                return jsonify({"error": str(e)}), 500

    return jsonify({"message": "Cliente non trovato"}), 404

# -------------------------------------------------------------------
#  H) GRAFICO RENDIMENTO
# -------------------------------------------------------------------
@app.route("/api/rendimento")
def rendimento_agenti():
    if "agente" not in session or session.get("ruolo", "").strip().lower() != "admin":
        return jsonify({"message": "Non autorizzato"}), 401

    mese = request.args.get("mese")  # può essere vuoto!
    anno = request.args.get("anno")

    if not anno:
        return jsonify({"message": "Anno obbligatorio"}), 400

    clienti = clienti_sheet.get_all_records(value_render_option='UNFORMATTED_VALUE')
    rendimento = {}

    print("\n📊 Diagnostica colonne Provvigione / Competenza")

    for cliente in clienti:
        provvigione = cliente.get("Provvigione")
        agente = cliente.get("Agente")
        comp_raw = cliente.get("Competenza", "")

        if not provvigione or not agente or not comp_raw:
            continue

        # ✅ Normalizzazione competenza
        if isinstance(comp_raw, (datetime, date)):
            competenza = comp_raw.strftime("%m/%Y")
        elif isinstance(comp_raw, int):
            base_date = date(1899, 12, 30)
            data_comp = base_date + timedelta(days=comp_raw)
            competenza = data_comp.strftime("%m/%Y")
        else:
            competenza = str(comp_raw).strip()

        # ↪️ Separiamo competenza in mese/anno
        if "/" not in competenza:
            continue

        mese_comp, anno_comp = competenza.split("/")

        if anno_comp != anno:
            continue

        if mese and mese_comp != mese:
            continue

        try:
            provvigione_float = float(str(provvigione).replace(",", "."))
        except:
            continue

        # ➕ Accumula provvigioni per agente e competenza
        if agente not in rendimento:
            rendimento[agente] = {}
        if competenza not in rendimento[agente]:
            rendimento[agente][competenza] = 0

        rendimento[agente][competenza] += provvigione_float

    # ➤ Output per frontend: rendimento[agente][competenza] = totale
    return jsonify(rendimento)



# 📊 Nuova route: rendimento per agente (solo dati propri)
@app.route("/api/rendimento-agente")
def rendimento_agente():
    if "agente" not in session:
        return jsonify({"message": "Non autenticato"}), 401

    agente_corrente = session['agente'].strip().upper()
    mese = request.args.get("mese")  # opzionale
    anno = request.args.get("anno")

    if not anno:
        return jsonify({"message": "Anno obbligatorio"}), 400

    clienti = clienti_sheet.get_all_records(value_render_option='UNFORMATTED_VALUE')
    rendimento = {}

    from datetime import datetime, date, timedelta

    for cliente in clienti:
        if str(cliente.get("Agente", "")).strip().upper() != agente_corrente:
            continue

        provvigione = cliente.get("Provvigione")
        comp_raw = cliente.get("Competenza", "")

        if isinstance(comp_raw, (datetime, date)):
            competenza = comp_raw.strftime("%m/%Y")
        elif isinstance(comp_raw, int):
            base_date = date(1899, 12, 30)
            data_comp = base_date + timedelta(days=comp_raw)
            competenza = data_comp.strftime("%m/%Y")
        else:
            competenza = str(comp_raw).strip()

        if not provvigione or not competenza or "/" not in competenza:
            continue

        mese_comp, anno_comp = competenza.split("/")
        if anno_comp != anno:
            continue

        if mese and mese_comp != mese:
            continue

        try:
            valore = float(str(provvigione).replace(",", "."))
        except:
            continue

        rendimento[competenza] = rendimento.get(competenza, 0) + valore

    # Ordina per competenza cronologica
    dati = [
        {"competenza": k, "totale": round(v, 2)}
        for k, v in sorted(rendimento.items(), key=lambda x: datetime.strptime(x[0], "%m/%Y"))
    ]

    return jsonify(dati)


@app.route("/rendimento-agente")
def pagina_rendimento_agente():
    if "agente" not in session:
        return redirect("/login")
    return render_template("rendimento_agente.html")



# route di accesso protetta

@app.route("/rendimento-agenti")
def pagina_rendimento_agenti():
    if 'agente' not in session or session.get('ruolo', '').lower() != 'admin':
        return redirect("/login")
    return render_template("rendimento.html")




###if __name__ == '__main__':
###    app.run(debug=True)


#@app.route("/")
#def serve_login():
#    return send_from_directory(app.static_folder, "login.html")

#@app.route("/index.html")
#def serve_index():
#    if 'agente' not in session:
#        return redirect("/")
#    return send_from_directory(app.static_folder, "index.html")

#import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
